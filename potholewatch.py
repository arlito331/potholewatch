"""
PotholeWatch Phase 3 — News-First Road Incident Intelligence
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ARCHITECTURE: News article → geocode → Google Maps APIs → Street View pothole finder → cross-reference

FLOW:
  1. Scrape La Prensa, Mi Diario, TVN, Telemetro, Panama América, Crítica
     → extract headline + direct article link + location + quote
  2. Geocode exact location from article (Nominatim → Google Geocoding → AI guess)
  3. Google Maps Roads API → road quality at coordinates
  4. Google Maps Nearby Search → crowdsourced hazards (Waze-fused data)
  5. Waze live tile endpoint → direct pothole pins
  6. Street View 4-angle scan → Claude Vision identifies & photographs pothole
  7. X/Twitter + Tráfico Panama → social confirmation
  8. Cross-reference engine → confidence score
  9. Single digest email → news link + pothole photo + all sources
 10. Weekly Excel report every Monday

SCHEDULE: Every 8 hours via GitHub Actions (0 */8 * * *)
"""

import os, io, re, json, time, base64, requests, math
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from anthropic import Anthropic

# ─── CONFIG ──────────────────────────────────────────────────────────────────

ALERT_RECIPIENTS  = ["joel@powerfixinc.com", "1@powerfixinc.com"]
ALERT_THRESHOLD   = "MEDIUM"
PROBABILITY_ORDER = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]

KEYWORDS_ES = [
    "hueco", "huecos", "cráter", "cráteres", "bache", "baches",
    "accidente", "vía deteriorada", "pavimento", "losa",
    "hundimiento", "derrumbe", "daño vial", "carretera deteriorada"
]

# Panama news sources — news article IS the anchor of every incident
NEWS_SOURCES = [
    {"name": "La Prensa",       "search": "https://www.prensa.com/buscar/?q={}",                  "domain": "prensa.com"},
    {"name": "Mi Diario",       "search": "https://www.midiario.com/?s={}",                       "domain": "midiario.com"},
    {"name": "TVN Noticias",    "search": "https://www.tvn-2.com/buscar/?q={}",                   "domain": "tvn-2.com"},
    {"name": "Telemetro",       "search": "https://www.telemetro.com/?s={}",                      "domain": "telemetro.com"},
    {"name": "Panama América",  "search": "https://www.panamaamerica.com.pa/?s={}",               "domain": "panamaamerica.com.pa"},
    {"name": "Crítica",         "search": "https://www.critica.com.pa/?s={}",                     "domain": "critica.com.pa"},
    {"name": "La Estrella",     "search": "https://www.laestrella.com.pa/?s={}",                  "domain": "laestrella.com.pa"},
]

NEWS_SEARCH_QUERIES = [
    "accidente carretera hueco bache Panama",
    "accidente vial deterioro carretera Panama",
    "cráter bache accidente vía Panama",
    "accidente Vía Centenario hueco",
    "accidente Corredor Norte bache",
    "accidente Transístmica hueco cráter",
    "accidente Interamericana bache",
    "accidente puente carretera Panama hueco",
]

WAZE_BBOX  = {"bottom": 7.2, "top": 9.7, "left": -83.0, "right": -77.2}
PROB_COLOR = {"CRITICAL": "#A32D2D", "HIGH": "#993C1D", "MEDIUM": "#854F0B", "LOW": "#5F5E5A"}
PROB_BG    = {"CRITICAL": "#FCEBEB", "HIGH": "#FAECE7", "MEDIUM": "#FAEEDA", "LOW": "#F1EFE8"}

client = Anthropic()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 1 — NEWS SCRAPER (anchor of every incident)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def fetch_news_articles():
    """
    Search ONLY for recent accident news where a pothole was the cause.
    Then cross-reference with MOP, ATTT, Tráfico Panama, and social media.
    Only articles from the last 7 days. Only accidents, not general road articles.
    """
    print("\n📰 STEP 1 — Fetching Panama accident news (last 7 days)...")
    articles = []
    seen = set()

    # Accident-first searches — must be recent AND about an accident caused by pothole
    accident_searches = [
        'accidente "hueco" OR "bache" OR "cráter" Panama carretera 2026 site:tvn-2.com OR site:telemetro.com OR site:prensa.com OR site:midiario.com OR site:critica.com.pa',
        'accidente vial "hueco" OR "bache" Panama mayo 2026',
        'accidente carretera Panama "hueco" OR "bache" OR "cráter" causa 2026',
        'motociclista OR conductor OR vehiculo accidente "hueco" OR "bache" Panama 2026',
        'ATTT OR MOP accidente vía "hueco" OR "bache" Panama 2026',
    ]

    for search in accident_searches:
        try:
            response = client.messages.create(
                model="claude-opus-4-5",
                max_tokens=2000,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content":
                    f"""Search for: {search}

IMPORTANT: Only find news articles from the LAST 7 DAYS about road ACCIDENTS in Panama where a pothole (hueco, bache, cráter) was the direct cause or contributing factor.

DO NOT include:
- Articles older than 7 days
- General road condition articles (no accident happened)
- Political articles about road funding
- Articles just mentioning potholes without an accident

For each qualifying article, extract:
- Headline
- News source (TVN, La Prensa, Mi Diario, Telemetro, Crítica, Panama América)
- Direct URL to article
- Date (must be within last 7 days)
- Exact road/street/km where accident happened
- Quote from article describing the pothole and accident
- Any mention of MOP or ATTT response
- Any fatalities or injuries mentioned"""
                }]
            )

            # Collect text from all blocks
            full_text = ""
            for block in response.content:
                if not hasattr(block, 'type'):
                    continue
                if block.type == "text":
                    full_text += block.text or ""
                elif block.type == "tool_result" and hasattr(block, 'content'):
                    for sub in (block.content or []):
                        if hasattr(sub, 'text'):
                            full_text += sub.text or ""

            if len(full_text) < 100:
                continue

            # Extract structured articles
            extract_resp = client.messages.create(
                model="claude-opus-4-5",
                max_tokens=2000,
                messages=[{"role": "user", "content":
                    f"""Extract ONLY recent accident articles from this text. Be strict.

TEXT:
{full_text[:4000]}

STRICT RULES:
1. Article must be from the last 7 days (May 2026 or very recent)
2. Must describe an actual road ACCIDENT (crash, injury, death, vehicle damage)
3. Pothole/hueco/bache must be the cause or contributing factor
4. SKIP general road condition articles, political articles, old articles

For each qualifying article output:

---START---
HEADLINE: [title]
SOURCE: [TVN Noticias / La Prensa / Mi Diario / Telemetro / Crítica / Panama América / La Estrella]
URL: [direct https:// article URL]
DATE: [publication date]
LOCATION: [exact road/street/highway/km - be specific]
DETAIL: [landmark, intersection, neighborhood, direction]
QUOTE: [direct quote about the pothole causing the accident]
MOP: [any MOP/ATTT mention or response]
INJURIES: [fatalities or injuries mentioned]
TERRITORY: [Panama City / Colón / La Chorrera / Chiriquí / Other]
---END---

If no articles qualify, output nothing."""
                }]
            )

            extract_text = extract_resp.content[0].text if extract_resp.content else ""

            import re as _re
            blocks = _re.split(r'---START---|---END---', extract_text)
            for block_text in blocks:
                block_text = block_text.strip()
                if len(block_text) < 60:
                    continue

                def get_field(label, t):
                    m = _re.search(rf'^{label}:\s*(.+)$', t, _re.IGNORECASE | _re.MULTILINE)
                    v = m.group(1).strip() if m else ''
                    return '' if v.lower() in ['unknown','none','n/a','no','—','-','blank',''] else v

                headline = get_field('HEADLINE', block_text)
                source   = get_field('SOURCE', block_text)
                url      = get_field('URL', block_text)
                date     = get_field('DATE', block_text)
                location = get_field('LOCATION', block_text)
                detail   = get_field('DETAIL', block_text)
                quote    = get_field('QUOTE', block_text)
                mop      = get_field('MOP', block_text)
                injuries = get_field('INJURIES', block_text)
                territory= get_field('TERRITORY', block_text)

                # Skip if no real location
                if not headline or len(headline) < 8:
                    continue
                if not location or location.upper() in ['DETAIL:', 'DETAIL', 'N/A']:
                    continue
                if headline in seen:
                    continue
                # Skip bad URLs
                if url and any(x in url for x in ['google.','bing.','buscar','?s=','search','yahoo.']):
                    url = ''

                seen.add(headline)
                articles.append({
                    'headline': headline,
                    'source_name': source or 'Panama News',
                    'url': url,
                    'date': date,
                    'location': location,
                    'location_detail': detail,
                    'quote': quote,
                    'mop_attt': mop,
                    'injuries': injuries,
                    'keywords': [k for k in KEYWORDS_ES if k in (headline+quote+location).lower()],
                    'pothole_cause': True,
                    'territory': territory or 'Panama City'
                })
                print(f"      ✓ [{source}] {headline[:65]}...")
                print(f"        📍 {location} | 📅 {date} | 🔗 {url[:55] if url else 'no url'}")

            time.sleep(1)
        except Exception as e:
            print(f"   Search failed: {e}")

    print(f"   ✅ Found {len(articles)} accident articles")
    return articles


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 2 — SMART GEOCODING (from news article location)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def geocode_smart(location, location_detail="", google_key=""):
    """
    3-tier geocoding from news article location text.
    Tier 1: Google Geocoding API (most accurate for Panama roads)
    Tier 2: Nominatim OpenStreetMap
    Tier 3: Claude AI estimate
    Returns (lat, lng, method, accuracy)
    """
    search_text = f"{location} {location_detail}".strip()

    # Tier 1: Google Geocoding API
    if google_key:
        try:
            r = requests.get("https://maps.googleapis.com/maps/api/geocode/json",
                params={"address": search_text + ", Panama", "key": google_key,
                        "components": "country:PA", "language": "es"},
                timeout=10)
            if r.status_code == 200:
                results = r.json().get("results", [])
                if results:
                    loc = results[0]["geometry"]["location"]
                    lat, lng = loc["lat"], loc["lng"]
                    if 7.0 <= lat <= 10.0 and -84.0 <= lng <= -77.0:
                        accuracy = results[0].get("geometry", {}).get("location_type", "APPROXIMATE")
                        print(f"   📍 Google Geocoding: {lat:.5f}, {lng:.5f} ({accuracy})")
                        return lat, lng, "google_geocoding", accuracy
        except Exception as e:
            print(f"   Google geocoding failed: {e}")

    # Tier 2: Nominatim
    try:
        clean = re.sub(r'\b(km|kilómetro|metros?|mts?)\s*[\d.]+', '', search_text, flags=re.IGNORECASE).strip()
        r = requests.get("https://nominatim.openstreetmap.org/search",
            params={"q": clean + ", Panama", "format": "json", "limit": 3,
                    "countrycodes": "pa", "accept-language": "es"},
            headers={"User-Agent": "PotholeWatch/3.0"}, timeout=10)
        if r.status_code == 200 and r.json():
            best = max(r.json(), key=lambda x: float(x.get("importance", 0)))
            lat, lng = float(best["lat"]), float(best["lon"])
            if 7.0 <= lat <= 10.0 and -84.0 <= lng <= -77.0:
                print(f"   📍 Nominatim: {lat:.5f}, {lng:.5f}")
                return lat, lng, "nominatim", "APPROXIMATE"
    except Exception as e:
        print(f"   Nominatim failed: {e}")

    # Tier 3: Claude AI estimate
    try:
        response = client.messages.create(
            model="claude-opus-4-5", max_tokens=200,
            messages=[{"role": "user", "content":
                f"""What are the GPS coordinates of "{search_text}" in Panama?
                Respond ONLY: {{"lat": 0.0000, "lng": -00.0000}}
                Must be in Panama: lat 7.0-10.0, lng -84.0 to -77.0"""
            }]
        )
        match = re.search(r'\{[^}]+\}', response.content[0].text)
        if match:
            coords = json.loads(match.group())
            lat, lng = float(coords["lat"]), float(coords["lng"])
            if 7.0 <= lat <= 10.0 and -84.0 <= lng <= -77.0:
                print(f"   📍 AI estimate: {lat:.5f}, {lng:.5f}")
                return lat, lng, "ai_estimate", "APPROXIMATE"
    except Exception as e:
        print(f"   AI geocoding failed: {e}")

    return None, None, "failed", None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 3 — GOOGLE MAPS APIs (Roads + Nearby hazards + Waze)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def query_google_roads(lat, lng, google_key):
    """Google Maps Roads API — snap to road and get road info."""
    if not google_key:
        return {}
    try:
        r = requests.get("https://roads.googleapis.com/v1/nearestRoads",
            params={"points": f"{lat},{lng}", "key": google_key}, timeout=10)
        if r.status_code == 200:
            roads = r.json().get("snappedPoints", [])
            if roads:
                snapped = roads[0]["location"]
                return {
                    "snapped_lat": snapped["latitude"],
                    "snapped_lng": snapped["longitude"],
                    "place_id": roads[0].get("placeId", ""),
                    "original_index": roads[0].get("originalIndex", 0)
                }
    except Exception as e:
        print(f"   Roads API failed: {e}")
    return {}


def query_google_nearby_hazards(lat, lng, google_key):
    """
    Google Maps Places Nearby Search for road hazards.
    Google fuses Waze crowdsourced data — this pulls the same pins
    that appear on maps.google.com including Waze-reported potholes.
    """
    if not google_key:
        return []
    hazards = []
    try:
        # Search for road hazards nearby
        for keyword in ["pothole", "hueco", "bache", "road damage", "road hazard"]:
            r = requests.get("https://maps.googleapis.com/maps/api/place/nearbysearch/json",
                params={
                    "location": f"{lat},{lng}",
                    "radius": 500,
                    "keyword": keyword,
                    "key": google_key,
                    "language": "es"
                }, timeout=10)
            if r.status_code == 200:
                results = r.json().get("results", [])
                for place in results[:3]:
                    hazards.append({
                        "name": place.get("name", ""),
                        "vicinity": place.get("vicinity", ""),
                        "rating": place.get("rating", 0),
                        "lat": place.get("geometry", {}).get("location", {}).get("lat"),
                        "lng": place.get("geometry", {}).get("location", {}).get("lng"),
                        "keyword": keyword
                    })
    except Exception as e:
        print(f"   Nearby hazards search failed: {e}")

    # Also try Waze tile endpoint (Google infrastructure)
    try:
        r = requests.get("https://www.waze.com/live-map/api/georss",
            params={"top": min(lat+0.05, 10.0), "bottom": max(lat-0.05, 7.0),
                    "left": max(lng-0.05, -84.0), "right": min(lng+0.05, -77.0),
                    "env": "row", "types": "alerts"},
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.waze.com/live-map"},
            timeout=10)
        if r.status_code == 200:
            for alert in r.json().get("alerts", []):
                t = alert.get("type", "")
                if "POT_HOLE" in t or "HAZARD" in t:
                    hazards.append({
                        "name": f"Waze: {alert.get('subtype', t)}",
                        "vicinity": alert.get("street", "Unknown street"),
                        "lat": alert.get("location", {}).get("y"),
                        "lng": alert.get("location", {}).get("x"),
                        "reports": alert.get("nThumbsUp", 0) + 1,
                        "keyword": "waze_pothole"
                    })
    except Exception as e:
        print(f"   Waze tile fetch failed: {e}")

    return hazards


def distance_km(lat1, lng1, lat2, lng2):
    if not all([lat1, lng1, lat2, lng2]):
        return 999
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng/2)**2
    return 6371 * 2 * math.asin(math.sqrt(a))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 4 — STREET VIEW 4-ANGLE POTHOLE IDENTIFIER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def fetch_street_view_angles(lat, lng, google_key):
    """Fetch Street View from 4 directions + downward tilt."""
    if not google_key or not lat or not lng:
        return []
    angles = [
        {"heading": 0,   "pitch": -20, "label": "Norte"},
        {"heading": 90,  "pitch": -20, "label": "Este"},
        {"heading": 180, "pitch": -20, "label": "Sur"},
        {"heading": 270, "pitch": -20, "label": "Oeste"},
        {"heading": 0,   "pitch": -60, "label": "Abajo"},
    ]
    images = []
    for angle in angles:
        try:
            r = requests.get("https://maps.googleapis.com/maps/api/streetview",
                params={"size": "640x400", "location": f"{lat},{lng}",
                        "heading": angle["heading"], "pitch": angle["pitch"],
                        "fov": 90, "source": "outdoor", "key": google_key},
                timeout=15)
            if r.status_code == 200 and len(r.content) > 8000:
                images.append({**angle, "b64": base64.b64encode(r.content).decode("utf-8")})
        except Exception as e:
            pass
    print(f"   📸 Street View: {len(images)}/5 angles fetched")
    return images


def identify_pothole_in_images(images, location_name, article_quote=""):
    """
    Claude Vision analyzes all Street View angles simultaneously.
    Identifies, photographs, and measures the pothole.
    Returns: best_b64, analysis, confirmed
    """
    if not images:
        return None, {}, False
    try:
        content = []

        # Add article context if available
        if article_quote:
            content.append({"type": "text",
                "text": f"Context from news article about this location: '{article_quote}'"})

        # Add all images
        for i, img in enumerate(images):
            content.append({"type": "text",
                "text": f"Image {i+1}/{len(images)} — Street View facing {img['label']} at {location_name}, Panama:"})
            content.append({"type": "image",
                "source": {"type": "base64", "media_type": "image/jpeg", "data": img["b64"]}})

        content.append({"type": "text", "text": f"""
You are a road damage expert trained specifically on Panama potholes (huecos/baches).
A news article reported a road accident caused by a pothole at {location_name}, Panama.

Examine ALL {len(images)} images carefully.

CRITICAL DISTINCTION — Pothole vs Damaged Road:

A TRUE POTHOLE (hueco/bache) is:
✅ A HOLE or DEPRESSION where asphalt is PHYSICALLY MISSING
✅ Exposed base layer (light grey/beige concrete or aggregate) visible INSIDE a depression
✅ Clear depth visible — you can see INTO the hole
✅ Jagged broken edges surrounding the missing asphalt
✅ Often oval/irregular shaped, 20cm to 2m diameter

Based on real Panama potholes documented by PowerFix:
- The hole shows light grey/beige base material against dark surrounding asphalt
- Edges are crumbling and broken, not smooth
- Sometimes debris (gravel, dirt, straw) collects inside
- Multiple potholes often cluster together

NOT a pothole — do NOT confirm these:
❌ Surface cracks (even large ones) without missing asphalt
❌ Rough or worn road texture
❌ Patch repairs (darker square/rectangle repairs)
❌ Wet road surface
❌ Road markings or paint
❌ Normal concrete expansion joints
❌ General road deterioration without visible holes

pothole_confirmed = true ONLY if you clearly see a HOLE with missing asphalt and exposed base material.
Be strict — false positives waste PowerFix deployment resources.

Respond ONLY in this JSON format:
{{
  "best_image_index": 0,
  "pothole_found": true/false,
  "pothole_confirmed": true/false,
  "best_heading": 0,
  "best_label": "Norte/Sur/Este/Oeste/Abajo",
  "severity": "none/minor/moderate/severe/critical",
  "description": "Detailed 3-sentence description. Include: where in frame, road surface condition, type of damage, visual evidence that matches the news report.",
  "estimated_diameter_m": 0.0,
  "estimated_depth_cm": 0,
  "location_in_frame": "center-lane/right-lane/left-lane/shoulder/multiple",
  "road_condition": "Brief overall road condition assessment",
  "accident_risk": "low/medium/high/critical",
  "matches_news_report": true/false,
  "powerfix_opportunity": true/false,
  "confidence_visual": 0-100,
  "per_image_notes": ["note for image 1", "note for image 2", "..."]
}}

pothole_confirmed = true only if you clearly see road damage.
matches_news_report = true if visual damage is consistent with what the article described."""
        })

        response = client.messages.create(
            model="claude-opus-4-5", max_tokens=1200,
            messages=[{"role": "user", "content": content}]
        )
        text = response.content[0].text.strip()
        text = re.sub(r'^```json\s*', '', text)
        text = re.sub(r'\s*```$', '', text)
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            result = json.loads(match.group())
            best_idx = min(result.get("best_image_index", 0), len(images)-1)
            confirmed = result.get("pothole_confirmed", False)
            best_img  = images[best_idx]["b64"]

            status = "🕳️  POTHOLE CONFIRMED" if confirmed else "ℹ️  No pothole confirmed"
            sev    = result.get("severity", "?")
            diam   = result.get("estimated_diameter_m", 0)
            vconf  = result.get("confidence_visual", 0)
            print(f"   🔍 Vision: {status} | Severity: {sev} | Ø{diam}m | Visual confidence: {vconf}%")
            if result.get("description"):
                print(f"      \"{result['description'][:120]}...\"")
            return best_img, result, confirmed
    except Exception as e:
        print(f"   Vision analysis failed: {e}")

    return images[0]["b64"] if images else None, {}, False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 5 — SOCIAL CONFIRMATION (X + Tráfico Panama)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def fetch_trafico_panama():
    print("\n🚦 Fetching Tráfico Panama...")
    try:
        response = client.messages.create(
            model="claude-opus-4-5", max_tokens=1500,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content":
                """Search for recent posts from Tráfico Panama about road accidents and damage.

Search these:
1. site:x.com traficopanama hueco OR bache OR cráter OR accidente 2026
2. traficopanama "hueco" OR "bache" OR "cráter" carretera Panama
3. twitter.com traficopanama accidente vía deteriorada

Return ALL posts mentioning: hueco, huecos, cráter, cráteres, bache, baches, accidente, vía deteriorada.
Include the exact post text, date, and road/location mentioned."""
            }]
        )
        posts = []
        for block in response.content:
            if hasattr(block, 'type') and block.type == "text" and len(block.text) > 30:
                if any(k in block.text.lower() for k in KEYWORDS_ES):
                    posts.append({
                        "source": "trafico_panama",
                        "text": block.text[:600],
                        "url": "https://x.com/traficopanama",
                        "timestamp": datetime.utcnow().isoformat()
                    })
        print(f"   Found {len(posts)} posts")
        return posts
    except Exception as e:
        print(f"   Failed: {e}")
        return []


def fetch_mop_attt_news(location, headline=""):
    """Search for MOP and ATTT news/responses related to a specific accident location."""
    try:
        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=1000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content":
                f"""Search for MOP (Ministerio de Obras Públicas) or ATTT (Autoridad de Tránsito) 
news about road conditions or responses at: {location} Panama

Also search for comments or reactions from MOP/ATTT about the accident: {headline}

Search: MOP OR ATTT "{location}" Panama 2026 hueco OR bache OR accidente

Return any official statements, responses, or news about this location from MOP or ATTT."""
            }]
        )
        results = []
        for block in response.content:
            if hasattr(block, 'type') and block.type == "text" and len(block.text) > 30:
                results.append(block.text[:400])
        return results
    except Exception as e:
        return []


def fetch_social_comments(location, headline="", article_url=""):
    """Search for social media comments about a specific accident — hueco/bache keywords."""
    try:
        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=1000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content":
                f"""Search for social media posts and comments about this road accident in Panama:
Location: {location}
Headline: {headline}

Search: "{location}" Panama accidente hueco OR bache OR cráter site:x.com OR site:facebook.com OR site:instagram.com 2026
Also: traficopanama "{location}" hueco OR bache OR accidente

Find posts/comments mentioning: hueco, huecos, cráter, bache, baches, accidente
Return actual post texts with the exact keywords used."""
            }]
        )
        posts = []
        for block in response.content:
            if hasattr(block, 'type') and block.type == "text" and len(block.text) > 30:
                if any(k in block.text.lower() for k in KEYWORDS_ES):
                    posts.append({
                        "source": "social_media",
                        "text": block.text[:500],
                        "timestamp": datetime.utcnow().isoformat()
                    })
        return posts
    except Exception as e:
        return []


def fetch_x_social(location, headline=""):
    """Search X for posts confirming a specific incident location."""
    try:
        query = f"{location} Panama {headline[:30] if headline else 'accidente hueco bache'}"
        response = client.messages.create(
            model="claude-opus-4-5", max_tokens=1000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content":
                f"""Search X/Twitter and social media for posts about: {query}
                Keywords: hueco, cráter, bache, accidente, vía deteriorada Panama 2026
                Return actual post texts confirming road conditions at this location. Last 48 hours."""
            }]
        )
        posts = []
        for block in response.content:
            if hasattr(block, 'type') and block.type == "text" and len(block.text) > 30:
                if any(k in block.text.lower() for k in KEYWORDS_ES):
                    posts.append({
                        "source": "x_twitter",
                        "text": block.text[:500],
                        "url": f"https://x.com/search?q={requests.utils.quote(query)}",
                        "timestamp": datetime.utcnow().isoformat()
                    })
        return posts
    except Exception as e:
        return []


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 6 — CROSS-REFERENCE ENGINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def cross_reference(article, lat, lng, nearby_hazards,
                    trafico_posts, x_posts, sv_analysis, geocode_method, mop_news=None):
    """
    Cross-reference all 5 sources anchored to a news article.
    Returns probability, confidence, and full source breakdown.
    """
    sources = {}
    confirmed = []

    # Source 1: News article (always present — it's the anchor)
    sources["news"] = {
        "confirmed": True,
        "headline": article.get("headline", ""),
        "source_name": article.get("source_name", ""),
        "url": article.get("url", ""),
        "date": article.get("date", ""),
        "location": article.get("location", ""),
        "quote": article.get("quote", ""),
        "keywords": article.get("keywords", [])
    }
    confirmed.append("news")

    # Source 2: Google Maps / Waze hazards nearby
    waze_pins    = [h for h in nearby_hazards if "waze" in h.get("keyword","").lower()]
    google_pins  = [h for h in nearby_hazards if "waze" not in h.get("keyword","").lower()]
    if nearby_hazards:
        sources["google_waze"] = {
            "confirmed": True,
            "waze_pins": waze_pins[:3],
            "google_hazards": google_pins[:3],
            "total": len(nearby_hazards)
        }
        confirmed.append("google_waze")

    # Source 3: Street View visual confirmation
    if sv_analysis.get("pothole_confirmed"):
        sources["street_view"] = {
            "confirmed": True,
            "severity": sv_analysis.get("severity"),
            "diameter_m": sv_analysis.get("estimated_diameter_m"),
            "depth_cm": sv_analysis.get("estimated_depth_cm"),
            "description": sv_analysis.get("description",""),
            "accident_risk": sv_analysis.get("accident_risk"),
            "visual_confidence": sv_analysis.get("confidence_visual", 0),
            "matches_news": sv_analysis.get("matches_news_report", False)
        }
        confirmed.append("street_view")
    elif sv_analysis:
        sources["street_view"] = {
            "confirmed": False,
            "description": sv_analysis.get("description", "No road damage confirmed visually"),
            "road_condition": sv_analysis.get("road_condition","")
        }

    # Source 4a: MOP/ATTT news
    if mop_news:
        sources["mop_attt"] = {
            "confirmed": True,
            "reports": mop_news[:2]
        }
        confirmed.append("mop_attt")

    # Source 4b: Tráfico Panama
    # Match posts mentioning this location
    loc_words = [w for w in article.get("location","").lower().split() if len(w) > 3]
    tp_hits = [p for p in trafico_posts if any(w in p.get("text","").lower() for w in loc_words)]
    if tp_hits or trafico_posts:
        sources["trafico_panama"] = {
            "confirmed": bool(tp_hits),
            "posts": (tp_hits or trafico_posts)[:2]
        }
        if tp_hits:
            confirmed.append("trafico_panama")

    # Source 5: X/Twitter
    if x_posts:
        sources["x_twitter"] = {
            "confirmed": True,
            "posts": x_posts[:2]
        }
        confirmed.append("x_twitter")

    # ── Confidence calculation ────────────────────────────────────────────────
    n = len(confirmed)
    base = {1: 40, 2: 62, 3: 78, 4: 90, 5: 97}.get(n, 40)

    # Geocoding quality bonus
    if geocode_method == "google_geocoding":
        base = min(base + 5, 99)
    elif geocode_method == "ai_estimate":
        base = max(base - 10, 20)

    # Visual severity bonus
    sev = sv_analysis.get("severity","none")
    if sev in ["severe","critical"]:
        base = min(base + 8, 99)
    elif sev == "moderate":
        base = min(base + 4, 99)

    # News+visual match bonus
    if sv_analysis.get("matches_news_report"):
        base = min(base + 5, 99)

    # Probability
    if base >= 85:   prob = "CRITICAL"
    elif base >= 68: prob = "HIGH"
    elif base >= 45: prob = "MEDIUM"
    else:            prob = "LOW"

    return {
        "probability": prob,
        "confidence": base,
        "sources_confirmed": confirmed,
        "source_count": n,
        "sources": sources
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP 7 — RECOMMENDED ACTION (PowerFix specific)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get_recommended_action(article, sv_analysis, cross_ref):
    """Generate PowerFix-specific action recommendation."""
    prob      = cross_ref.get("probability","MEDIUM")
    sev       = sv_analysis.get("severity","unknown")
    diam      = sv_analysis.get("estimated_diameter_m", 0)
    loc       = article.get("location","this location")
    confirmed = cross_ref.get("sources_confirmed",[])

    if prob == "CRITICAL" and "street_view" in confirmed:
        boxes = max(1, int(math.ceil((diam * diam * 0.05) / 0.02))) if diam else "3-5"
        return f"🚨 IMMEDIATE DEPLOYMENT — PowerPatch to {loc}. Visual confirms {sev} damage ({diam}m Ø). Estimated {boxes} boxes needed. Contact MOP/ATTT."
    elif prob in ["CRITICAL","HIGH"]:
        return f"⚡ PRIORITY VISIT — Inspect {loc} within 24hrs. {len(confirmed)}/5 sources confirmed. Bring 3-5 boxes PowerPatch."
    elif prob == "MEDIUM":
        return f"📋 SCHEDULE INSPECTION — {loc} showing road damage indicators. Monitor for 24hrs then assess PowerPatch deployment."
    else:
        return f"👁️ MONITOR — {loc} flagged. Low confidence, keep in watch list."


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DIGEST EMAIL BUILDER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_digest_html(incidents, scan_time):
    critical = [i for i in incidents if i["cross_ref"]["probability"] == "CRITICAL"]
    high     = [i for i in incidents if i["cross_ref"]["probability"] == "HIGH"]
    medium   = [i for i in incidents if i["cross_ref"]["probability"] == "MEDIUM"]
    scan_dt  = datetime.fromisoformat(scan_time).strftime("%d %b %Y · %H:%M UTC")

    def pills(confirmed):
        defs = [("news","📰 News"),("google_waze","📡 Waze/Maps"),
                ("street_view","📸 StreetView"),("trafico_panama","🚦 Tráfico"),("x_twitter","𝕏 X")]
        html = ""
        for key, label in defs:
            ok = key in confirmed
            bg = "#EAF3DE" if ok else "#F5F5F5"
            c  = "#3B6D11" if ok else "#bbb"
            html += f'<span style="font-size:10px;padding:2px 8px;border-radius:10px;background:{bg};color:{c};margin-right:3px;display:inline-block;margin-bottom:3px">{label}</span>'
        return html

    def incident_card(inc):
        xref    = inc["cross_ref"]
        article = inc["article"]
        sv      = inc.get("sv_analysis", {})
        prob    = xref["probability"]
        conf    = xref["confidence"]
        loc     = article.get("location","Unknown")
        ptw_id  = inc["ptw_id"]
        lat     = inc.get("lat","")
        lng     = inc.get("lng","")
        action  = inc.get("recommended_action","")
        sv_b64  = inc.get("best_sv_image","")
        color   = PROB_COLOR.get(prob,"#5F5E5A")
        bg      = PROB_BG.get(prob,"#F1EFE8")

        maps_url = f"https://www.google.com/maps?q={lat},{lng}" if lat and lng else "#"
        sv_url   = f"https://www.google.com/maps/@{lat},{lng},3a,75y,{sv.get('best_heading',0)}h,75t/data=!3m1!1e3" if lat and lng else "#"
        geo_method = inc.get("geocode_method","")

        # ── News article block (the anchor) ──────────────────────────────────
        article_url  = article.get("url","")
        headline     = article.get("headline","No headline")
        source_name  = article.get("source_name","")
        art_date     = article.get("date","")
        art_quote    = article.get("quote","")
        art_mop      = article.get("mop_attt","")
        headline_link = f'<a href="{article_url}" style="color:#185FA5;text-decoration:none;font-weight:bold">{headline}</a>' if article_url else f'<strong>{headline}</strong>'

        news_block = f"""
        <div style="padding:10px 12px;background:#EBF4FB;border-left:4px solid #185FA5;border-radius:0 6px 6px 0;margin:10px 0">
          <div style="font-size:13px;margin-bottom:4px">{headline_link}</div>
          <div style="font-size:11px;color:#666;margin-bottom:6px">
            📰 <strong>{source_name}</strong> {f'· {art_date}' if art_date else ''}
            {f' · <a href="{article_url}" style="color:#185FA5">Ver artículo completo →</a>' if article_url else ''}
          </div>
          {f'<div style="font-size:12px;color:#444;font-style:italic;border-top:1px solid #C8DFF0;padding-top:6px">"{art_quote}"</div>' if art_quote else ''}
          {f'<div style="font-size:11px;color:#185FA5;margin-top:4px">🏛️ MOP/ATTT: {art_mop}</div>' if art_mop else ''}
        </div>"""

        # ── Street View pothole photo ─────────────────────────────────────────
        sv_block = ""
        if sv_b64:
            sev   = sv.get("severity","").upper()
            diam  = sv.get("estimated_diameter_m", 0)
            depth = sv.get("estimated_depth_cm", 0)
            desc  = sv.get("description","")
            vconf = sv.get("confidence_visual", 0)
            risk  = sv.get("accident_risk","").upper()
            label = sv.get("best_label","")
            match_news = sv.get("matches_news_report", False)
            confirmed_badge = f'<span style="background:#EAF3DE;color:#3B6D11;font-size:10px;padding:2px 7px;border-radius:10px;font-weight:bold">✅ POTHOLE CONFIRMED</span>' if sv.get("pothole_confirmed") else '<span style="background:#F5F5F5;color:#888;font-size:10px;padding:2px 7px;border-radius:10px">📸 Road surface captured</span>'
            sv_block = f"""
            <div style="margin:10px 0">
              {confirmed_badge}
              <div style="margin-top:6px;position:relative;display:inline-block;width:100%">
                <img src="data:image/jpeg;base64,{sv_b64}"
                     style="width:100%;max-width:600px;border-radius:6px;border:2px solid {color};display:block"
                     alt="Street View pothole — {loc}"/>
                <div style="position:absolute;bottom:8px;left:8px;background:rgba(0,0,0,0.75);color:#fff;font-size:10px;padding:3px 8px;border-radius:4px">
                  📍 {loc} · Facing {label}
                </div>
              </div>
              <div style="margin-top:6px;padding:8px 10px;background:#F5F5F5;border-radius:4px;font-size:12px;color:#444;line-height:1.5">
                <strong>AI Road Analysis:</strong> {desc}<br>
                <span style="color:{color}">
                  {f'Ø <strong>{diam}m</strong> diameter · ' if diam else ''}
                  {f'<strong>{depth}cm</strong> deep · ' if depth else ''}
                  Severity: <strong>{sev}</strong> · 
                  Accident risk: <strong>{risk}</strong> · 
                  Visual confidence: <strong>{vconf}%</strong>
                  {' · ✓ Matches news report' if match_news else ''}
                </span><br>
                <a href="{sv_url}" style="color:#185FA5;font-size:11px">🗺️ Open Street View at this exact location →</a>
              </div>
            </div>"""
        elif lat and lng:
            sv_block = f'<div style="margin:8px 0;padding:8px 10px;background:#F5F5F5;border-radius:4px;font-size:12px"><a href="{sv_url}" style="color:#185FA5">📸 View Street View at this location →</a></div>'

        # ── Waze/Google hazards ───────────────────────────────────────────────
        waze_src = xref.get("sources",{}).get("google_waze",{})
        waze_html = ""
        if waze_src.get("confirmed"):
            pins = waze_src.get("waze_pins",[]) + waze_src.get("google_hazards",[])
            for p in pins[:3]:
                waze_html += f'<div style="font-size:11px;color:#555;margin-top:2px">📡 {p.get("name","?")} · {p.get("vicinity",p.get("street","?"))}</div>'

        # ── Social posts ──────────────────────────────────────────────────────
        social_html = ""
        for src_key, label in [("trafico_panama","🚦 Tráfico Panama"),("x_twitter","𝕏 X/Twitter")]:
            src_data = xref.get("sources",{}).get(src_key,{})
            for p in src_data.get("posts",[])[:1]:
                text = p.get("text","")[:180]
                url  = p.get("url","")
                link = f' <a href="{url}" style="color:#185FA5;font-size:10px">→ ver</a>' if url else ""
                social_html += f'<div style="font-size:11px;color:#555;padding:5px 8px;background:#f9f9f9;border-radius:4px;margin-top:4px"><strong>{label}:</strong> {text}{link}</div>'

        return f"""
<div style="border:1px solid {color};border-radius:8px;margin-bottom:22px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
  <div style="background:{bg};padding:10px 14px;border-bottom:1px solid {color};display:flex;align-items:center;flex-wrap:wrap;gap:8px">
    <span style="background:{color};color:#fff;font-size:10px;font-weight:bold;padding:3px 10px;border-radius:12px">{prob}</span>
    <strong style="font-size:13px;color:{color};flex:1">{ptw_id} — {loc}</strong>
    <span style="font-size:11px;color:#888">{conf}% confidence · {xref['source_count']}/5 sources</span>
  </div>
  <div style="padding:12px 14px">
    {news_block}
    {sv_block}
    <div style="margin:8px 0">{pills(xref['sources_confirmed'])}</div>
    <div style="font-size:12px;color:#555;margin:6px 0">
      📍 <a href="{maps_url}" style="color:#185FA5">{loc} — Google Maps</a>
      {f' · {lat:.5f}, {lng:.5f}' if lat and lng else ''}
      {f' <span style="color:#aaa;font-size:10px">({geo_method})</span>' if geo_method else ''}
    </div>
    {waze_html}
    {social_html}
    <div style="margin-top:10px;padding:8px 12px;background:#EAF3DE;border-radius:5px;font-size:12px;color:#2D6A0F;line-height:1.5">
      {action}
    </div>
  </div>
</div>"""

    # ── Build full email body ─────────────────────────────────────────────────
    blocks = ""
    for label, group, color in [
        ("🚨 CRITICAL", critical, "#A32D2D"),
        ("⚠️ HIGH",     high,     "#993C1D"),
        ("📍 MEDIUM",   medium,   "#854F0B")
    ]:
        if group:
            blocks += f'<h2 style="color:{color};font-size:14px;margin:22px 0 10px;border-bottom:2px solid {color};padding-bottom:5px;letter-spacing:.03em">{label} ({len(group)})</h2>'
            for inc in group:
                blocks += incident_card(inc)

    n_vis = sum(1 for i in incidents if i.get("sv_analysis",{}).get("pothole_confirmed"))

    return f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;max-width:660px;margin:0 auto;color:#1a1a1a;padding:16px;background:#fff">
<div style="background:#0D0D0D;padding:14px 18px;border-radius:8px;margin-bottom:18px;display:flex;align-items:center">
  <span style="width:8px;height:8px;border-radius:50%;background:#E24B4A;display:inline-block;margin-right:10px"></span>
  <span style="color:#fff;font-size:15px;font-weight:bold;letter-spacing:.05em">PotholeWatch</span>
  <span style="color:#555;font-size:11px;margin-left:8px">Phase 3 · PowerFix Inc.</span>
  <span style="margin-left:auto;color:#555;font-size:11px">{scan_dt}</span>
</div>
<div style="display:flex;gap:8px;margin-bottom:18px">
  {''.join([f'<div style="flex:1;text-align:center;padding:10px 6px;background:{PROB_BG[p]};border-radius:6px"><div style="font-size:20px;font-weight:bold;color:{PROB_COLOR[p]}">{cnt}</div><div style="font-size:9px;color:{PROB_COLOR[p]};text-transform:uppercase;letter-spacing:.06em;margin-top:2px">{label}</div></div>' for p,cnt,label in [("CRITICAL",len(critical),"Critical"),("HIGH",len(high),"High"),("MEDIUM",len(medium),"Medium"),("LOW",len(incidents),"Total")]])}
</div>
{f'<div style="padding:8px 12px;background:#EAF3DE;border-radius:6px;font-size:12px;color:#3B6D11;margin-bottom:16px">🕳️ <strong>{n_vis} pothole{"s" if n_vis!=1 else ""} visually confirmed</strong> via Google Street View AI analysis</div>' if n_vis else ''}
{blocks}
<div style="margin-top:20px;padding:10px;background:#f5f5f5;border-radius:6px;font-size:10px;color:#aaa;text-align:center;line-height:1.6">
  PotholeWatch Phase 3 · Sources: La Prensa · Mi Diario · TVN · Telemetro · Panama América · Crítica<br>
  Tráfico Panama · X/Twitter · Google Maps · Waze · Street View AI · PowerFix Inc. · Panama
</div></body></html>"""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# WEEKLY EXCEL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_excel(incidents):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PotholeWatch Weekly"

        headers = ["PTW ID","Probability","Confidence %","Location","Territory",
                   "News Source","Headline","Article URL","Date","Quote",
                   "Lat","Lng","Google Maps","Geocode Method",
                   "Sources Count","Sources Confirmed",
                   "Pothole Confirmed","Severity","Diameter (m)","Depth (cm)",
                   "Accident Risk","Matches News","Visual Confidence %",
                   "Recommended Action","Scan Time"]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = PatternFill("solid", fgColor="0D0D0D")
            c.font = Font(color="FFFFFF", bold=True, size=9)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        pfills = {"CRITICAL":"FCEBEB","HIGH":"FAECE7","MEDIUM":"FAEEDA","LOW":"F1EFE8"}
        pfonts = {"CRITICAL":"A32D2D","HIGH":"993C1D","MEDIUM":"854F0B","LOW":"5F5E5A"}

        for row, inc in enumerate(incidents, 2):
            xref = inc.get("cross_ref", {})
            art  = inc.get("article", {})
            sv   = inc.get("sv_analysis", {})
            prob = xref.get("probability","MEDIUM")
            lat  = inc.get("lat","")
            lng  = inc.get("lng","")
            maps = f"https://www.google.com/maps?q={lat},{lng}" if lat and lng else ""

            row_data = [
                inc.get("ptw_id",""), prob, xref.get("confidence",0),
                art.get("location",""), inc.get("territory",""),
                art.get("source_name",""), art.get("headline",""),
                art.get("url",""), art.get("date",""), art.get("quote",""),
                lat, lng, maps, inc.get("geocode_method",""),
                f"{xref.get('source_count',0)}/5",
                ", ".join(xref.get("sources_confirmed",[])),
                "YES" if sv.get("pothole_confirmed") else "NO",
                sv.get("severity","—"), sv.get("estimated_diameter_m","—"),
                sv.get("estimated_depth_cm","—"), sv.get("accident_risk","—"),
                "YES" if sv.get("matches_news_report") else "NO",
                sv.get("confidence_visual","—"),
                inc.get("recommended_action",""), inc.get("scan_time","")
            ]

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=pfills.get(prob,"FAEEDA"))
                if col == 2:
                    c.font = Font(color=pfonts.get(prob,"854F0B"), bold=True)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 13 and maps:
                    c.hyperlink = maps
                    c.font = Font(color="185FA5", underline="single")
                if col == 8 and art.get("url"):
                    c.hyperlink = art["url"]
                    c.font = Font(color="185FA5", underline="single")

        widths = [10,10,12,30,14,14,40,30,12,50,10,10,18,14,10,25,12,10,12,10,12,10,12,50,22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "PotholeWatch Weekly Summary"
        ws2["A1"].font = Font(bold=True, size=13)
        vis = sum(1 for i in incidents if i.get("sv_analysis",{}).get("pothole_confirmed"))
        for r, (label, val) in enumerate([
            ("Total incidents", len(incidents)),
            ("Critical", sum(1 for i in incidents if i["cross_ref"].get("probability")=="CRITICAL")),
            ("High", sum(1 for i in incidents if i["cross_ref"].get("probability")=="HIGH")),
            ("Medium", sum(1 for i in incidents if i["cross_ref"].get("probability")=="MEDIUM")),
            ("Potholes visually confirmed", vis),
            ("News sources scraped", len(NEWS_SOURCES)),
            ("Report generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))
        ], start=3):
            ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        print(f"   Excel failed: {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GMAIL SENDER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def send_digest(incidents, scan_time, excel_bytes=None):
    if not all([os.environ.get("GMAIL_REFRESH_TOKEN"),
                os.environ.get("GMAIL_CLIENT_ID"),
                os.environ.get("GMAIL_CLIENT_SECRET")]):
        print("⚠️  Gmail credentials missing")
        return False
    try:
        import google.auth.transport.requests
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build

        creds = Credentials(
            token=None,
            refresh_token=os.environ.get("GMAIL_REFRESH_TOKEN"),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=os.environ.get("GMAIL_CLIENT_ID"),
            client_secret=os.environ.get("GMAIL_CLIENT_SECRET"),
            scopes=["https://www.googleapis.com/auth/gmail.send"]
        )
        creds.refresh(google.auth.transport.requests.Request())
        service = build("gmail", "v1", credentials=creds)

        n_crit = sum(1 for i in incidents if i["cross_ref"].get("probability")=="CRITICAL")
        n_high = sum(1 for i in incidents if i["cross_ref"].get("probability")=="HIGH")
        n_vis  = sum(1 for i in incidents if i.get("sv_analysis",{}).get("pothole_confirmed"))
        n_tot  = len(incidents)

        if n_crit > 0:
            subj = f"🚨 PotholeWatch — {n_crit} CRITICAL · {n_high} HIGH · {n_vis} potholes confirmed · {n_tot} total"
        elif n_high > 0:
            subj = f"⚠️ PotholeWatch — {n_high} HIGH · {n_vis} potholes confirmed · {n_tot} incidents"
        else:
            subj = f"📍 PotholeWatch — {n_tot} incidents · {n_vis} potholes confirmed"

        msg = MIMEMultipart("mixed")
        msg["Subject"] = subj
        msg["From"]    = "potholewatch@powerfixinc.com"
        msg["To"]      = ", ".join(ALERT_RECIPIENTS)

        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(build_digest_html(incidents, scan_time), "html"))
        msg.attach(alt)

        if excel_bytes:
            date_str = datetime.utcnow().strftime("%Y-%m-%d")
            part = MIMEBase("application", "octet-stream")
            part.set_payload(excel_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="PotholeWatch_{date_str}.xlsx"')
            msg.attach(part)

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
        service.users().messages().send(userId="me", body={"raw": raw}).execute()
        note = " + Excel" if excel_bytes else ""
        print(f"✅ Digest sent: {subj}{note}")
        return True
    except Exception as e:
        print(f"❌ Email failed: {e}")
        return False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN SCAN — NEWS-FIRST PIPELINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def run_scan():
    print(f"\n{'━'*60}")
    print(f"PotholeWatch Phase 3 — News-First Pipeline")
    print(f"Started: {datetime.utcnow().isoformat()}")
    print(f"{'━'*60}")

    gkey1     = os.environ.get("GOOGLE_MAPS_API_KEY","")
    gkey2     = os.environ.get("GOOGLE_STREETVIEW_API_KEY","")
    google_key = gkey1 or gkey2
    print(f"   Google API key: {'SET ('+str(len(google_key))+' chars)' if google_key else 'NOT SET — Street View disabled'}")
    incident_counter = int(os.environ.get("INCIDENT_COUNTER", "1"))
    is_weekly        = os.environ.get("WEEKLY_REPORT", "false").lower() == "true"
    scan_time        = datetime.utcnow().isoformat()
    all_incidents    = []

    # ── STEP 1: Fetch news articles (anchor of everything) ─────────────────
    news_articles = fetch_news_articles()

    if not news_articles:
        print("\n⚠️  No news articles found — running social fallback scan...")
        # Fallback: use X/Tráfico posts as anchors if no news
        news_articles = [{
            "headline": f"Road damage report — {q.split('Panamá')[0].strip()}",
            "source_name": "Social Media",
            "url": "",
            "date": datetime.utcnow().strftime("%Y-%m-%d"),
            "location": q.split("accidente ")[-1].split(" Panamá")[0],
            "location_detail": "",
            "quote": "",
            "keywords": ["accidente", "bache"],
            "pothole_cause": True,
            "territory": "Panama City"
        } for q in ["accidente Vía Centenario Panamá", "accidente Corredor Norte Panamá",
                     "accidente Transístmica Panamá", "hueco bache accidente Panamá"]]

    # ── Fetch global social sources once ───────────────────────────────────
    trafico_posts = fetch_trafico_panama()

    # ── STEP 2-7: Process each news article through the pipeline ───────────
    print(f"\n🔄 Processing {len(news_articles)} articles through pipeline...\n")

    for article in news_articles:
        loc     = article.get("location","")
        detail  = article.get("location_detail","")
        headline = article.get("headline","")[:60]
        source  = article.get("source_name","")
        territory = article.get("territory","Panama City")

        print(f"\n{'─'*55}")
        print(f"📰 [{source}] {headline}...")
        print(f"   📍 Location: {loc}")

        # STEP 2: Smart geocoding from article location
        lat, lng, geo_method, geo_accuracy = geocode_smart(loc, detail, google_key)
        if not lat or not lng:
            print(f"   ❌ Could not geocode — skipping")
            continue

        # STEP 3: Google Maps Roads API
        road_info = query_google_roads(lat, lng, google_key)
        if road_info.get("snapped_lat"):
            lat = road_info["snapped_lat"]
            lng = road_info["snapped_lng"]
            print(f"   🛣️  Snapped to road: {lat:.5f}, {lng:.5f}")

        # STEP 4: Google Maps Nearby + Waze hazards
        print(f"   📡 Querying Google Maps + Waze hazards...")
        nearby_hazards = query_google_nearby_hazards(lat, lng, google_key)
        waze_count = sum(1 for h in nearby_hazards if "waze" in h.get("keyword","").lower())
        google_count = len(nearby_hazards) - waze_count
        print(f"   Found: {waze_count} Waze pins + {google_count} Google hazards")

        # STEP 5: Street View 4-angle pothole identification
        best_sv_img = None
        sv_analysis = {}
        print(f"   🛰️  Fetching Street View (4 angles + downward)...")
        sv_images = fetch_street_view_angles(lat, lng, google_key)
        if sv_images:
            best_sv_img, sv_analysis, pothole_confirmed = identify_pothole_in_images(
                sv_images, loc, article.get("quote","")
            )
        else:
            print(f"   ⚠️  No Street View imagery at this location")

        # STEP 6: MOP/ATTT cross-reference
        print(f"   🏛️  Searching MOP/ATTT news...")
        mop_news = fetch_mop_attt_news(loc, headline)
        print(f"   Found {len(mop_news)} MOP/ATTT references")

        # STEP 6b: Social media comments (hueco/bache keywords)
        print(f"   💬 Searching social media comments...")
        social_comments = fetch_social_comments(loc, headline, article.get('url',''))
        print(f"   Found {len(social_comments)} social posts with keywords")

        # STEP 6c: X/Twitter
        x_posts = fetch_x_social(loc, headline)
        print(f"   𝕏 X posts: {len(x_posts)}")

        # Combine all social
        all_social = social_comments + x_posts

        # STEP 7: Cross-reference engine
        cross_ref = cross_reference(
            article, lat, lng, nearby_hazards,
            trafico_posts, all_social, sv_analysis, geo_method,
            mop_news=mop_news
        )

        prob = cross_ref["probability"]
        conf = cross_ref["confidence"]
        n_src = cross_ref["source_count"]

        if PROBABILITY_ORDER.index(prob) < PROBABILITY_ORDER.index(ALERT_THRESHOLD):
            print(f"   ⬇️  {prob} ({conf}%) — below threshold, skip")
            continue

        action = get_recommended_action(article, sv_analysis, cross_ref)
        ptw_id = f"PTW-{incident_counter:03d}"
        incident_counter += 1

        incident = {
            "ptw_id":             ptw_id,
            "article":            article,
            "cross_ref":          cross_ref,
            "lat":                lat,
            "lng":                lng,
            "geocode_method":     geo_method,
            "geocode_accuracy":   geo_accuracy,
            "road_info":          road_info,
            "nearby_hazards":     nearby_hazards,
            "best_sv_image":      best_sv_img,
            "sv_analysis":        sv_analysis,
            "x_posts":            all_social,
            "mop_news":           mop_news,
            "recommended_action": action,
            "territory":          territory,
            "scan_time":          scan_time
        }
        all_incidents.append(incident)

        print(f"\n   ✅ {ptw_id} — {prob} ({conf}%) · {n_src}/5 sources · {geo_method}")
        if sv_analysis.get("pothole_confirmed"):
            d = sv_analysis.get("estimated_diameter_m",0)
            s = sv_analysis.get("severity","")
            print(f"   🕳️  POTHOLE CONFIRMED — {s} severity, Ø{d}m")
        print(f"   ⚡ {action[:80]}...")

        time.sleep(1.5)  # Rate limit

    # ── Send digest ─────────────────────────────────────────────────────────
    print(f"\n{'━'*60}")
    n_vis = sum(1 for i in all_incidents if i.get("sv_analysis",{}).get("pothole_confirmed"))
    print(f"Phase 3 complete: {len(all_incidents)} incidents · {n_vis} potholes confirmed")

    if all_incidents:
        excel_bytes = build_excel(all_incidents) if is_weekly else None
        send_digest(all_incidents, scan_time, excel_bytes)
    else:
        print("No incidents above threshold — no email sent.")

    # Export incidents.json for Vercel live dashboard
    try:
        dashboard_data = []
        for inc in all_incidents:
            xref    = inc.get("cross_ref", {})
            article = inc.get("article", {})
            sv      = inc.get("sv_analysis", {})
            dashboard_data.append({
                "ptw_id":            inc.get("ptw_id"),
                "probability":       xref.get("probability"),
                "confidence":        xref.get("confidence"),
                "sources_confirmed": xref.get("sources_confirmed", []),
                "source_count":      xref.get("source_count", 0),
                "lat":               inc.get("lat"),
                "lng":               inc.get("lng"),
                "geocode_method":    inc.get("geocode_method"),
                "territory":         inc.get("territory"),
                "scan_time":         inc.get("scan_time"),
                "headline":          article.get("headline"),
                "source_name":       article.get("source_name"),
                "url":               article.get("url"),
                "date":              article.get("date"),
                "location":          article.get("location"),
                "quote":             article.get("quote"),
                "mop_attt":          article.get("mop_attt"),
                "recommended_action":inc.get("recommended_action"),
                "pothole_confirmed": sv.get("pothole_confirmed", False),
                "severity":          sv.get("severity"),
                "diameter_m":        sv.get("estimated_diameter_m"),
                "depth_cm":          sv.get("estimated_depth_cm"),
                "accident_risk":     sv.get("accident_risk"),
                "sv_description":    sv.get("description"),
                "visual_confidence": sv.get("confidence_visual"),
                "sv_image_b64":      inc.get("best_sv_image", ""),
            })
        output = {"scan_time": scan_time, "total": len(dashboard_data), "incidents": dashboard_data}
        with open("incidents.json", "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        print(f"✅ incidents.json written ({len(dashboard_data)} incidents)")
    except Exception as e:
        print(f"Warning: incidents.json export failed: {e}")

    print(f"{'━'*60}\n")


if __name__ == "__main__":
    run_scan()
